import tkinter as tk
from tkinter import simpledialog, filedialog, messagebox
from tkinter import ttk
from datetime import datetime
import openpyxl

class SpendingTracker(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Rastreador de Gastos")
        self.geometry("500x300")
        self.configure(bg="#a3d2a5")

        self.despesas = []

        self.mes_var = tk.StringVar()
        self.ano_var = tk.StringVar()

        self.combo_mes = ttk.Combobox(self, textvariable=self.mes_var, state="readonly")
        self.combo_mes["values"] = [
            "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro",
            "Novembro", "Dezembro",
        ]
        self.combo_mes.current(datetime.now().month - 1)
        self.combo_mes.pack(pady=5)

        self.combo_ano = ttk.Combobox(self, textvariable=self.ano_var, state="readonly")
        self.combo_ano["values"] = [str(ano) for ano in range(2022, datetime.now().year + 2)]
        self.combo_ano.current(datetime.now().year - 2022)
        self.combo_ano.pack(pady=5)

        self.frame_entrada = tk.Frame(self, bg="#a3d2a5")
        self.frame_entrada.pack(pady=10)

        self.rotulo_despesa = tk.Label(self.frame_entrada, text="Despesa:", bg="#a3d2a5")
        self.rotulo_despesa.grid(row=0, column=0, padx=5, pady=5)
        self.entrada_despesa = tk.Entry(self.frame_entrada, width=20)
        self.entrada_despesa.grid(row=0, column=1, padx=5, pady=5)

        self.rotulo_valor = tk.Label(self.frame_entrada, text="Valor:", bg="#a3d2a5")
        self.rotulo_valor.grid(row=1, column=0, padx=5, pady=5)
        self.entrada_valor = tk.Entry(self.frame_entrada, width=10)
        self.entrada_valor.grid(row=1, column=1, padx=5, pady=5)

        self.rotulo_vencimento = tk.Label(self.frame_entrada, text="Vencimento:", bg="#a3d2a5")
        self.rotulo_vencimento.grid(row=2, column=0, padx=5, pady=5)
        self.entrada_vencimento = tk.Entry(self.frame_entrada, width=10)
        self.entrada_vencimento.grid(row=2, column=1, padx=5, pady=5)

        self.rotulo_parcelas = tk.Label(self.frame_entrada, text="Parcelas Acordo:", bg="#a3d2a5")
        self.rotulo_parcelas.grid(row=3, column=0, padx=5, pady=5)
        self.entrada_parcelas = tk.Entry(self.frame_entrada, width=7)
        self.entrada_parcelas.grid(row=3, column=1, padx=5, pady=5)

        self.rotulo_observacao = tk.Label(self.frame_entrada, text="Observação:", bg="#a3d2a5")
        self.rotulo_observacao.grid(row=4, column=0, padx=5, pady=5)
        self.entrada_observacao = tk.Entry(self.frame_entrada, width=20)
        self.entrada_observacao.grid(row=4, column=1, padx=5, pady=5)

        self.botao_adicionar = tk.Button(self.frame_entrada, text="Adicionar Despesa",
                                         command=self.dialogo_adicionar_despesa)
        self.botao_adicionar.grid(row=5, column=0, columnspan=2, padx=5, pady=5)

        self.botao_excluir = tk.Button(self, text="Excluir Despesa", command=self.excluir_despesa)
        self.botao_excluir.pack(pady=5)

        self.botao_salvar_excel = tk.Button(self, text="Salvar no Excel", command=self.salvar_no_excel)
        self.botao_salvar_excel.pack(pady=10)

        self.botao_sair = tk.Button(self, text="Sair", command=self.destroy)
        self.botao_sair.pack(pady=10)

    def dialogo_adicionar_despesa(self):
        despesa = self.entrada_despesa.get()
        valor = self.entrada_valor.get()
        vencimento = self.entrada_vencimento.get()
        parcelas = self.entrada_parcelas.get()
        observacao = self.entrada_observacao.get()

        if not despesa or not valor or not vencimento:
            messagebox.showerror("Erro", "Campos obrigatórios não preenchidos.")
            return

        try:
            valor = float(valor.replace(',', '.'))  # Substitui vírgula por ponto e converte para float
            if parcelas:
                parcelas = int(parcelas)
        except ValueError:
            messagebox.showerror("Erro", "Valor e Parcelas devem ser números.")
            return

        parcelas_acordo = f"{parcelas}x" if parcelas else ""

        self.despesas.append(
            (self.combo_ano.get(), self.combo_mes.get(), despesa, valor, vencimento, parcelas_acordo, observacao))

        self.entrada_despesa.delete(0, "end")
        self.entrada_valor.delete(0, "end")
        self.entrada_vencimento.delete(0, "end")
        self.entrada_parcelas.delete(0, "end")
        self.entrada_observacao.delete(0, "end")

    def excluir_despesa(self):
        arquivo_existente = filedialog.askopenfilename(
            title="Escolher arquivo existente",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )

        if arquivo_existente:
            self.carregar_despesas(arquivo_existente)
            despesa_selecionada = simpledialog.askinteger("Excluir Despesa",
                                                          "Digite o número da despesa a ser excluída:", minvalue=1,
                                                          maxvalue=len(self.despesas))

            if despesa_selecionada is not None:
                index = despesa_selecionada - 1
                resposta = messagebox.askyesno("Confirmar Exclusão", "Deseja realmente excluir esta despesa?")
                if resposta:
                    del self.despesas[index]

    def carregar_despesas(self, arquivo_existente):
        try:
            planilha = openpyxl.load_workbook(arquivo_existente)
            folha = planilha.active
            self.despesas = [tuple(row) for row in folha.iter_rows(values_only=True)][1:]
            planilha.close()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar despesas do arquivo: {str(e)}")

    def salvar_no_excel(self):
        if not self.despesas:
            messagebox.showinfo("Informação", "Nenhuma despesa para salvar.")
            return

        caminho_arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            title="Salvar despesas no Excel",
        )

        if not caminho_arquivo:
            return

        planilha = openpyxl.Workbook()
        folha = planilha.active
        folha.append(["Ano", "Mês", "Despesa", "Valor", "Vencimento", "Parcelas Acordo", "Observação"])

        for despesa in self.despesas:
            folha.append(despesa)

        for row_num in range(2, folha.max_row + 1):
            folha[f'D{row_num}'].number_format = 'R$#,##0.00'

        for row in folha.iter_rows(min_row=2, max_row=folha.max_row, max_col=folha.max_column):
            for cell in row:
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                     right=openpyxl.styles.Side(style='thin'),
                                                     top=openpyxl.styles.Side(style='thin'),
                                                     bottom=openpyxl.styles.Side(style='thin'))

        # Adiciona célula "Total" na coluna O (15ª coluna)
        total_valor = sum(despesa[3] for despesa in self.despesas)
        folha.append(["Total", f'R${total_valor:,.2f}'])
        folha[f'O{folha.max_row}'].number_format = 'R$#,##0.00'  # Formatação R$ para a célula "Total"

        planilha.save(caminho_arquivo)
        messagebox.showinfo("Informação", "Despesas salvas no Excel.")

if __name__ == "__main__":
    app = SpendingTracker()
    app.mainloop()
